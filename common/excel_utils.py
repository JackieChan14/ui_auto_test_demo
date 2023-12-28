# _*_coding:utf-8_*_
"""
@ Author:陈金泉
@ Date:2023/9/5 14:37
@ Description:
"""
import os.path

import openpyxl


class NewExcelReader:
	def __init__(self):
		self.workbook = None
		self.sheet = None
		self.rows = 0
		self.read_rows = 0  # 当前读取到的行数
	
	def open_excel(self, file_path):
		if not os.path.isfile(file_path):
			raise FileNotFoundError('%s is not a file' % file_path)
		
		openpyxl.Workbook.encoding = 'utf-8'
		self.workbook = openpyxl.load_workbook(filename=file_path)
		self.sheet = self.workbook[self.workbook.sheetnames[0]]
		self.rows = self.sheet.max_row
		self.read_rows = 0
	
	def get_sheets(self):
		sheets = self.workbook.sheetnames
		return sheets
	
	def set_sheets(self, name):
		self.sheet = self.workbook[name]
		self.rows = self.sheet.max_row
		self.read_rows = 0
	
	def read_line(self) -> list:
		lines = list()
		for row in self.sheet.rows:
			line = list()
			for cell in row:
				if cell.value is None:
					line.append('')
				else:
					line.append(cell.value)
			lines.append(line)
		return lines


class OldExelReader:
	pass
